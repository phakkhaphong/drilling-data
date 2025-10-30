#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor: Seam Codes from DH70.xlsx (all systems available in worksheet)
Outputs a DataFrame with columns:
  seam_id, system_id, system_name, seam_label, seam_code, priority, description, created_at
"""

import pandas as pd
import openpyxl
from datetime import datetime
from typing import Optional


def extract_seam_codes(excel_path: str = "data/raw/DH70.xlsx") -> pd.DataFrame:
	wb = openpyxl.load_workbook(excel_path, data_only=True)
	ws = wb['Seam Code']

	# Column positions based on actual worksheet structure (1-based displayed here):
	# [30, 'Seam Label', 'Seam Code', 46, 'Seam Label', 'Seam Code', ...]
	systems = [
		{'id': '30', 'name': 'System_30', 'label_col': 2, 'code_col': 3, 'priority': 6},
		{'id': '46', 'name': 'System_46', 'label_col': 5, 'code_col': 6, 'priority': 5},
		{'id': '57', 'name': 'System_57', 'label_col': 8, 'code_col': 9, 'priority': 4},
		{'id': '58', 'name': 'System_58', 'label_col': 11, 'code_col': 12, 'priority': 3},
		{'id': 'Quality', 'name': 'Quality_System', 'label_col': 14, 'code_col': 15, 'priority': 1},
		{'id': '73', 'name': 'System_73', 'label_col': 17, 'code_col': 18, 'priority': 2},
	]

	rows = []
	seam_id = 1
	for system in systems:
		for r in range(2, ws.max_row + 1):
			label = ws.cell(r, system['label_col']).value
			code = ws.cell(r, system['code_col']).value
			if label and code is not None:
				try:
					code_val: Optional[int] = int(code) if isinstance(code, (int, float)) else None
					if code_val is not None:
						rows.append({
							'seam_id': seam_id,
							'system_id': system['id'],
							'system_name': system['name'],
							'seam_label': str(label).strip(),
							'seam_code': code_val,
							'priority': system['priority'],
							'description': f"{system['name']} - {str(label).strip()}",
							'created_at': datetime.now(),
						})
						seam_id += 1
				except Exception:
					pass

	wb.close()
	return pd.DataFrame(rows)


__all__ = ["extract_seam_codes"]

