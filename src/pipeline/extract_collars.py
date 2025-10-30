#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor: Collars from DH70.xlsx (DAT201 worksheet)
Outputs a DataFrame with columns:
  collar_id, hole_id, easting, northing, elevation, final_depth,
  dip, drilling_date, azimuth, contractor, remarks, created_at, updated_at
"""

import pandas as pd
import openpyxl
from datetime import datetime
from typing import Dict, Tuple


def extract_collars(excel_path: str = "data/raw/DH70.xlsx") -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['DAT201']

    # Load as DataFrame using first row as header
    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append(row)
    wb.close()
    df = pd.DataFrame(data[1:], columns=data[0])

    # Map first occurrence row per hole to header cell values from that row
    # Columns: B(1)=easting, C(2)=northing, D(3)=elevation, G(6)=total_depth, Z(25)=year_drilled,
    #           AA(26)=Geologist, AH(33)=DH_Version, AG(32)=Block No
    hole_first_row_info: Dict[str, dict] = {}
    seen_holes = set()
    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None for v in row):
            continue
        # Assume DHID is in a cell of the row; find it by header name or value position
        # Try common positions by headerized DataFrame: if row aligns with data rows, 'DHID' will be in the same column index
        # Fallback: if the first non-empty token looks like a hole id (e.g., 'BC01C'), use it
        dhid_value = None
        try:
            # Find by column name if available in header
            if 'DHID' in data[0]:
                idx = data[0].index('DHID')
                if idx < len(row):
                    dhid_value = row[idx]
        except Exception:
            dhid_value = None
        if dhid_value is None:
            # Heuristic: first non-empty string token
            for cell in row:
                if isinstance(cell, str) and cell.strip():
                    dhid_value = cell.strip()
                    break
        if not dhid_value or not isinstance(dhid_value, str):
            continue
        hole_id_guess = dhid_value.strip()
        if hole_id_guess not in seen_holes:
            easting_bc = row[1] if len(row) > 1 else None  # B
            northing_bc = row[2] if len(row) > 2 else None  # C
            elevation_d = row[3] if len(row) > 3 else None  # D
            total_depth_g = row[6] if len(row) > 6 else None  # G
            year_drilled_z = row[25] if len(row) > 25 else None  # Z
            geologist_aa = row[26] if len(row) > 26 else None  # AA
            dh_version_ah = row[33] if len(row) > 33 else None  # AH
            block_no_ag = row[32] if len(row) > 32 else None  # AG
            hole_first_row_info[hole_id_guess] = {
                'easting': easting_bc,
                'northing': northing_bc,
                'elevation': elevation_d,
                'total_depth': total_depth_g,
                'year_drilled': year_drilled_z,
                'geologist': geologist_aa,
                'dh_version': dh_version_ah,
                'block_no': block_no_ag,
            }
            seen_holes.add(hole_id_guess)

    # Unique holes
    unique_holes = df['DHID'].dropna().unique()
    rows = []
    collar_id = 1
    for hole_id in unique_holes:
        hole_df = df[df['DHID'] == hole_id]
        first = hole_df.iloc[0]
        # We take total_depth from header G; we will not use max depth of intervals
        max_depth = None

        def pick_first(cols):
            for c in cols:
                if c in hole_df.columns:
                    val = hole_df[c].dropna().head(1)
                    if len(val) > 0:
                        return val.iloc[0]
            return None

        # Try direct column names first
        azimuth = pick_first(['Azimuth', 'AZIMUTH', 'Azi', 'AZI'])
        dip = pick_first(['Dip', 'DIP', 'Inclination', 'Incl.'])
        drilling_date = pick_first(['Drilling Date', 'Date', 'Drill Date'])
        contractor = pick_first(['Contractor', 'Drilling Contractor', 'Contr.'])
        remarks = pick_first(['Remarks', 'Remark', 'Comments', 'Notes'])
        elevation_val = pick_first(['Elevation', 'RL', 'Reduced Level'])
        geologist_val = pick_first(['Geologist', 'Geo', 'GEO'])
        block_no_val = pick_first(['Block', 'Block No', 'Block_no', 'BlockNo'])
        dh_version_val = pick_first(['DH Version', 'Version', 'DH_Version'])

        # Fallback: scan first few header-like rows for key tokens (values spread across columns)
        if any(v is None for v in [azimuth, dip, drilling_date, contractor, remarks, elevation_val, geologist_val, block_no_val, dh_version_val]):
            scan_rows = hole_df.head(6).fillna("")
            for _, row in scan_rows.iterrows():
                cells = [str(x).strip() for x in row.tolist() if str(x).strip() != ""]
                for i, cell in enumerate(cells):
                    low = cell.lower()
                    def next_val():
                        return cells[i+1] if i + 1 < len(cells) else None
                    if azimuth is None and ('azimuth' in low or low in ('azi','az')):
                        azimuth = next_val()
                    if dip is None and ('dip' in low or 'incl' in low):
                        dip = next_val()
                    if drilling_date is None and ('date' in low):
                        drilling_date = next_val()
                    if contractor is None and ('contractor' in low or 'drill' in low and 'contract' in low):
                        contractor = next_val()
                    if remarks is None and ('remark' in low or 'note' in low or 'comment' in low):
                        remarks = next_val()
                    if elevation_val is None and (low == 'rl' or 'elevation' in low or 'reduced level' in low):
                        elevation_val = next_val()
                    if geologist_val is None and ('geologist' in low or low in ('geo','geol')):
                        geologist_val = next_val()
                    if block_no_val is None and ('block' in low):
                        block_no_val = next_val()
                    if dh_version_val is None and ('version' in low):
                        dh_version_val = next_val()

            # Positional fallback for Block No at AG (index 32) within the first few rows
            if block_no_val is None:
                for _, r2 in scan_rows.iterrows():
                    lst = list(r2.values.tolist())
                    if len(lst) > 32:
                        v = lst[32]
                        if v is not None and str(v).strip() != "":
                            block_no_val = str(v).strip()
                            break

        # Normalize types
        def to_float(x):
            try:
                return float(x)
            except Exception:
                return None
        def to_date(x):
            if isinstance(x, datetime):
                return x.date()
            if isinstance(x, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(x, fmt).date()
                    except Exception:
                        continue
            return None

        # Fallback to B/C from header-mapped positions if direct columns are missing
        if (first.get('Easting') is None or first.get('Northing') is None) and str(hole_id).strip() in hole_first_row_info:
            info = hole_first_row_info[str(hole_id).strip()]
            if first.get('Easting') is None:
                e_val = to_float(info.get('easting'))
            else:
                e_val = to_float(first['Easting'])
            if first.get('Northing') is None:
                n_val = to_float(info.get('northing'))
            else:
                n_val = to_float(first['Northing'])
        else:
            e_val = to_float(first['Easting']) if 'Easting' in first and first['Easting'] is not None else None
            n_val = to_float(first['Northing']) if 'Northing' in first and first['Northing'] is not None else None

        rows.append({
            'collar_id': collar_id,
            'hole_id': str(hole_id).strip(),
            'easting': e_val,
            'northing': n_val,
            'elevation': (
                to_float(first['Elevation']) if 'Elevation' in first and first['Elevation'] is not None
                else to_float(elevation_val) if elevation_val is not None
                else (
                    to_float(hole_first_row_info[str(hole_id).strip()]['elevation']) if str(hole_id).strip() in hole_first_row_info and hole_first_row_info[str(hole_id).strip()].get('elevation') is not None else None
                )
            ),
            'total_depth': (
                to_float(hole_first_row_info[str(hole_id).strip()]['total_depth']) if str(hole_id).strip() in hole_first_row_info and hole_first_row_info[str(hole_id).strip()].get('total_depth') is not None else None
            ),
            'dip': to_float(dip) if dip is not None else None,
            'year_drilled': (
                int(hole_first_row_info[str(hole_id).strip()]['year_drilled']) if str(hole_id).strip() in hole_first_row_info and hole_first_row_info[str(hole_id).strip()].get('year_drilled') not in (None, "") else None
            ),
            # We no longer populate final_depth/drilling_date per requirement
            'azimuth': to_float(azimuth) if azimuth is not None else None,
            'contractor': None if contractor is None else str(contractor),
            'geologist': (
                str(geologist_val) if geologist_val is not None else (
                    str(hole_first_row_info[str(hole_id).strip()]['geologist']) if str(hole_id).strip() in hole_first_row_info and hole_first_row_info[str(hole_id).strip()].get('geologist') is not None else None
                )
            ),
            'block_no': (
                str(block_no_val) if block_no_val is not None else (
                    str(hole_first_row_info[str(hole_id).strip()]['block_no']) if str(hole_id).strip() in hole_first_row_info and hole_first_row_info[str(hole_id).strip()].get('block_no') is not None else None
                )
            ),
            'dh_version': (
                to_float(dh_version_val) if dh_version_val is not None else (
                    to_float(hole_first_row_info[str(hole_id).strip()]['dh_version']) if str(hole_id).strip() in hole_first_row_info and hole_first_row_info[str(hole_id).strip()].get('dh_version') is not None else None
                )
            ),
            'remarks': None if remarks is None else str(remarks),
            'created_at': datetime.now(),
            'updated_at': datetime.now(),
        })
        collar_id += 1

    return pd.DataFrame(rows)


__all__ = ["extract_collars"]

