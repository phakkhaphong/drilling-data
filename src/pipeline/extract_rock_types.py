#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor: Rock Types from DH70.xlsx (Rock Code worksheet)
Outputs a DataFrame with columns:
  rock_code, lithology, detail, created_at
- Sorted by rock_code ascending
- rock_code is the surrogate key used in relationships
"""

import pandas as pd
import openpyxl
from datetime import datetime
from typing import Optional


def extract_rock_types(excel_path: str = "data/raw/DH70.xlsx") -> pd.DataFrame:
	wb = openpyxl.load_workbook(excel_path, data_only=True)
	ws = wb['Rock Code']

	rows = []
	for r in range(2, ws.max_row + 1):
		detail = ws.cell(r, 1).value
		lithology = ws.cell(r, 2).value
		rock_code = ws.cell(r, 3).value
		if rock_code is not None and lithology is not None:
			try:
				code_val: Optional[int] = int(rock_code) if isinstance(rock_code, (int, float)) else None
				if code_val is not None:
					rows.append({
						'rock_code': code_val,
						'lithology': str(lithology).strip(),
						'detail': str(detail).strip() if detail else '',
						'created_at': datetime.now(),
					})
			except Exception:
				pass

	wb.close()

	df = pd.DataFrame(rows)
	# Sort by rock_code ascending
	df = df.sort_values(by=['rock_code'], kind='stable').reset_index(drop=True)
	# Ensure rock_code is integer
	df['rock_code'] = df['rock_code'].astype('Int64')
	return df


__all__ = ["extract_rock_types"]
