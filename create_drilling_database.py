#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Coal Drilling Database Generator
Creates drilling database with proper foreign key relationships
Following Best Practices for Mining/Drilling Data Management
"""

import polars as pl
import os
from datetime import datetime
import openpyxl

def create_drilling_database():
    """Create drilling database following coal drilling best practices"""
    
    print("="*70)
    print("CREATING COAL DRILLING DATABASE")
    print("Following Best Practices for Mining Data Management")
    print("="*70)
    
    # Load lookup data from Excel
    print("\n1. Loading lookup data from DH70.xlsx...")
    
    # Read seam codes from Excel
    wb = openpyxl.load_workbook('data/raw/DH70.xlsx', data_only=True)
    ws_seam = wb['Seam Code']
    
    # Extract seam codes from all systems
    seam_data = []
    systems = ['30', '46', '57', '58', 'Quality', '73']
    seam_id = 1
    
    for system in systems:
        # Find system column
        header_row = 1
        system_col = None
        for col in range(1, ws_seam.max_column + 1):
            cell_value = ws_seam.cell(header_row, col).value
            if cell_value and str(system) in str(cell_value):
                system_col = col
                break
        
        if system_col:
            # Read seam codes for this system
            row = header_row + 1
            while row <= ws_seam.max_row:
                seam_code = ws_seam.cell(row, system_col).value
                seam_label = ws_seam.cell(row, system_col + 1).value if system_col + 1 <= ws_seam.max_column else None
                
                if seam_code is not None and seam_label is not None:
                    try:
                        seam_code_val = int(seam_code) if isinstance(seam_code, (int, float)) else None
                        if seam_code_val is not None:
                            seam_data.append({
                                'seam_id': seam_id,
                                'system_id': system,
                                'system_name': f'System_{system}' if system.isdigit() else f'{system}_System',
                                'seam_label': str(seam_label).strip(),
                                'seam_code': seam_code_val
                            })
                            seam_id += 1
                    except:
                        pass
                row += 1
    
    seam_codes = pl.DataFrame(seam_data)
    
    # Read rock codes from Excel
    ws_rock = wb['Rock Code']
    rock_data = []
    rock_id = 1
    
    for row in range(2, ws_rock.max_row + 1):
        rock_code = ws_rock.cell(row, 1).value
        lithology = ws_rock.cell(row, 2).value
        detail = ws_rock.cell(row, 3).value
        
        if rock_code is not None:
            try:
                rock_code_val = int(rock_code) if isinstance(rock_code, (int, float)) else None
                if rock_code_val is not None and lithology:
                    rock_data.append({
                        'rock_id': rock_id,
                        'rock_code': rock_code_val,
                        'lithology': str(lithology).strip(),
                        'detail': str(detail).strip() if detail else ''
                    })
                    rock_id += 1
            except:
                pass
    
    rock_codes = pl.DataFrame(rock_data)
    wb.close()
    
    # Convert system_id to string for consistent comparison
    seam_codes = seam_codes.with_columns([
        pl.col('system_id').cast(pl.Utf8)
    ])
    
    print(f"Loaded {seam_codes.height} seam codes from {seam_codes['system_id'].n_unique()} systems")
    print(f"Loaded {rock_codes.height} rock codes")
    
    # Load existing data from database or SQLite
    print("\n2. Loading existing data...")
    
    # Try to load from SQLite database first
    if os.path.exists('drilling_database.db'):
        import sqlite3
        conn = sqlite3.connect('drilling_database.db')
        sample_analyses = pl.read_database('SELECT * FROM sample_analyses', conn)
        collars = pl.read_database('SELECT * FROM collars', conn)
        lithology_logs = pl.read_database('SELECT * FROM lithology_logs', conn)
        rock_types = pl.read_database('SELECT * FROM rock_types', conn)
        conn.close()
    else:
        # Load from existing drilling_database folder
        sample_analyses = pl.read_csv('data/drilling_database/sample_analyses.csv', infer_schema_length=10000)
        collars = pl.read_csv('data/drilling_database/collars.csv')
        lithology_logs = pl.read_csv('data/drilling_database/lithology_logs.csv')
        rock_types = pl.read_csv('data/drilling_database/rock_types.csv')
    
    print(f"Loaded {sample_analyses.height} sample analyses")
    print(f"Loaded {collars.height} collars")
    print(f"Loaded {lithology_logs.height} lithology logs")
    print(f"Loaded {rock_types.height} rock types")
    
    # Create drilling database with foreign keys
    print("\n3. Creating drilling database with foreign keys...")
    
    # Map seam codes to their IDs
    quality_seam_mapping = (seam_codes
                           .filter(pl.col('system_id') == 'Quality')
                           .select(['seam_id', 'seam_code'])
                           .with_columns([
                               pl.col('seam_code').cast(pl.Float64)
                           ])
                           .rename({'seam_id': 'seam_quality_id'}))
    
    seam_73_mapping = (seam_codes
                      .filter(pl.col('system_id') == '73')
                      .select(['seam_id', 'seam_code'])
                      .with_columns([
                          pl.col('seam_code').cast(pl.Float64)
                      ])
                      .rename({'seam_id': 'seam_73_id'}))
    
    # Create sample analyses with foreign keys
    sample_analyses_db = (sample_analyses
                        .join(quality_seam_mapping, 
                              left_on='seam_code_quality', 
                              right_on='seam_code', 
                              how='left')
                        .join(seam_73_mapping, 
                              left_on='seam_code_73', 
                              right_on='seam_code', 
                              how='left', 
                              suffix='_73')
                        .drop(['quality_seam_label', 'seam_label_73'])
                        .rename({'seam_code_quality': 'seam_code_quality_original'}))
    
    # Create output directory
    os.makedirs('data/drilling_database', exist_ok=True)
    
    # Save drilling database files
    print("\n4. Saving drilling database files...")
    sample_analyses_db.write_csv('data/drilling_database/sample_analyses.csv')
    seam_codes.write_csv('data/drilling_database/seam_codes_lookup.csv')
    rock_codes.write_csv('data/drilling_database/rock_codes_lookup.csv')
    
    # Copy other tables
    collars.write_csv('data/drilling_database/collars.csv')
    lithology_logs.write_csv('data/drilling_database/lithology_logs.csv')
    rock_types.write_csv('data/drilling_database/rock_types.csv')
    
    print("Saved all normalized tables")
    
    # Create summary statistics
    print("\n5. Creating summary statistics...")
    samples_with_quality = sample_analyses_db.filter(pl.col('seam_quality_id').is_not_null()).height
    samples_with_73 = sample_analyses_db.filter(pl.col('seam_73_id').is_not_null()).height
    
    print(f"\nSamples with Quality seam codes: {samples_with_quality}")
    print(f"Samples with 73 seam codes: {samples_with_73}")
    print(f"Total samples: {sample_analyses_db.height}")
    
    return sample_analyses_db, seam_codes, rock_codes, collars, lithology_logs, rock_types

def create_drilling_database_schema():
    """Create SQL schema following best practices for coal drilling"""
    
    print("\n6. Creating drilling database schema...")
    
    sql_content = """-- =====================================================
-- Coal Drilling Database Schema
-- Following Best Practices for Mining/Drilling Data Management
-- =====================================================
-- Based on: DH70.xlsx lookup data
-- Created: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """
-- =====================================================

-- Drop existing tables if they exist (in correct order due to foreign keys)
DROP TABLE IF EXISTS sample_analyses_normalized;
DROP TABLE IF EXISTS lithology_logs;
DROP TABLE IF EXISTS collars;
DROP TABLE IF EXISTS seam_codes_lookup;
DROP TABLE IF EXISTS rock_codes_lookup;
DROP TABLE IF EXISTS rock_types;

-- =====================================================
-- 1. LOOKUP TABLES
-- =====================================================

-- Seam Codes Lookup Table
-- Contains all seam codes from different classification systems
CREATE TABLE seam_codes_lookup (
    seam_id INTEGER PRIMARY KEY,
    system_id VARCHAR(20) NOT NULL,
    system_name VARCHAR(50) NOT NULL,
    seam_label VARCHAR(50) NOT NULL,
    seam_code INTEGER NOT NULL,
    description TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(system_id, seam_label, seam_code)
);

-- Rock Codes Lookup Table
-- Contains standard rock/lithology codes
CREATE TABLE rock_codes_lookup (
    rock_id INTEGER PRIMARY KEY,
    rock_code INTEGER UNIQUE NOT NULL,
    lithology VARCHAR(20) NOT NULL,
    detail VARCHAR(100) NOT NULL,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

-- Rock Types Lookup Table
CREATE TABLE rock_types (
    rock_type_id INTEGER PRIMARY KEY,
    rock_code INTEGER NOT NULL,
    rock_type VARCHAR(50) NOT NULL,
    FOREIGN KEY (rock_code) REFERENCES rock_codes_lookup(rock_code)
);

-- =====================================================
-- 2. CORE DRILLING TABLES
-- =====================================================

-- Collars Table
-- Contains drilling hole collar information (location, elevation, etc.)
CREATE TABLE collars (
    collar_id INTEGER PRIMARY KEY,
    hole_id VARCHAR(50) UNIQUE NOT NULL,
    easting REAL,
    northing REAL,
    elevation REAL,
    azimuth REAL,
    dip REAL,
    final_depth REAL,
    drilling_date DATE,
    contractor VARCHAR(100),
    remarks TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

-- Lithology Logs Table
-- Contains detailed lithology information for each depth interval
CREATE TABLE lithology_logs (
    log_id INTEGER PRIMARY KEY,
    hole_id VARCHAR(50) NOT NULL,
    depth_from REAL NOT NULL,
    depth_to REAL NOT NULL,
    rock_code INTEGER,
    description TEXT,
    FOREIGN KEY (hole_id) REFERENCES collars(hole_id),
    FOREIGN KEY (rock_code) REFERENCES rock_codes_lookup(rock_code),
    CHECK (depth_to > depth_from)
);

-- Sample Analyses Table
-- Contains laboratory analysis results for coal samples
CREATE TABLE sample_analyses (
    sample_id INTEGER PRIMARY KEY,
    hole_id VARCHAR(50) NOT NULL,
    depth_from REAL NOT NULL,
    depth_to REAL NOT NULL,
    sample_no VARCHAR(50) NOT NULL,
    
    -- Proximate Analysis
    im REAL,                    -- Inherent Moisture (%)
    tm REAL,                    -- Total Moisture (%)
    ash REAL,                   -- Ash Content (%)
    vm REAL,                   -- Volatile Matter (%)
    fc REAL,                   -- Fixed Carbon (%)
    
    -- Ultimate Analysis
    sulphur REAL,              -- Sulphur Content (%)
    
    -- Calorific Value
    gross_cv REAL,             -- Gross Calorific Value (kcal/kg)
    net_cv REAL,               -- Net Calorific Value (kcal/kg)
    
    -- Physical Properties
    sg REAL,                   -- Specific Gravity
    rd REAL,                   -- Relative Density
    hgi REAL,                 -- Hardgrove Grindability Index
    
    -- Seam Classification (Foreign Keys)
    seam_code_quality_original REAL,  -- Original quality seam code (for reference)
    seam_quality_id INTEGER,         -- Foreign key to seam_codes_lookup (Quality system)
    seam_73_id INTEGER,              -- Foreign key to seam_codes_lookup (73 system)
    
    -- Metadata
    analysis_date DATE,
    lab_name VARCHAR(100),
    remarks TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    -- Foreign key constraints
    FOREIGN KEY (hole_id) REFERENCES collars(hole_id),
    FOREIGN KEY (seam_quality_id) REFERENCES seam_codes_lookup(seam_id),
    FOREIGN KEY (seam_73_id) REFERENCES seam_codes_lookup(seam_id),
    
    -- Data validation constraints
    CHECK (depth_to > depth_from),
    CHECK (ash >= 0 AND ash <= 100),
    CHECK (vm >= 0 AND vm <= 100),
    CHECK (fc >= 0 AND fc <= 100),
    CHECK (im >= 0 AND im <= 100),
    CHECK (tm >= 0 AND tm <= 100)
);

-- =====================================================
-- 3. INDEXES FOR PERFORMANCE
-- =====================================================

-- Collars indexes
CREATE INDEX idx_collars_hole_id ON collars(hole_id);
CREATE INDEX idx_collars_location ON collars(easting, northing);

-- Lithology logs indexes
CREATE INDEX idx_lithology_hole_id ON lithology_logs(hole_id);
CREATE INDEX idx_lithology_depth ON lithology_logs(hole_id, depth_from, depth_to);
CREATE INDEX idx_lithology_rock_code ON lithology_logs(rock_code);

-- Sample analyses indexes
CREATE INDEX idx_sample_analyses_hole_id ON sample_analyses(hole_id);
CREATE INDEX idx_sample_analyses_depth ON sample_analyses(hole_id, depth_from, depth_to);
CREATE INDEX idx_sample_analyses_seam_quality ON sample_analyses(seam_quality_id);
CREATE INDEX idx_sample_analyses_seam_73 ON sample_analyses(seam_73_id);
CREATE INDEX idx_sample_analyses_sample_no ON sample_analyses(hole_id, sample_no);

-- Lookup tables indexes
CREATE INDEX idx_seam_codes_system ON seam_codes_lookup(system_id);
CREATE INDEX idx_seam_codes_code ON seam_codes_lookup(seam_code);
CREATE INDEX idx_seam_codes_label ON seam_codes_lookup(seam_label);
CREATE INDEX idx_rock_codes_code ON rock_codes_lookup(rock_code);
CREATE INDEX idx_rock_codes_lithology ON rock_codes_lookup(lithology);

-- =====================================================
-- 4. VIEWS FOR EASIER QUERYING
-- =====================================================

-- View: Complete Sample Analyses with Seam Information
CREATE VIEW sample_analyses_complete AS
SELECT 
    sa.*,
    sqc.seam_label as quality_seam_label,
    sqc.system_name as quality_system_name,
    s73c.seam_label as seam_73_label,
    s73c.system_name as seam_73_system_name
FROM sample_analyses sa
LEFT JOIN seam_codes_lookup sqc ON sa.seam_quality_id = sqc.seam_id
LEFT JOIN seam_codes_lookup s73c ON sa.seam_73_id = s73c.seam_id;

-- View: Sample Analyses with Collar Information
CREATE VIEW sample_analyses_with_location AS
SELECT 
    sa.*,
    c.easting,
    c.northing,
    c.elevation,
    c.final_depth as hole_final_depth
FROM sample_analyses sa
LEFT JOIN collars c ON sa.hole_id = c.hole_id;

-- View: Complete Hole Information
CREATE VIEW holes_complete AS
SELECT 
    c.*,
    COUNT(DISTINCT sa.sample_id) as sample_count,
    COUNT(DISTINCT ll.log_id) as log_count,
    MIN(sa.depth_from) as min_sample_depth,
    MAX(sa.depth_to) as max_sample_depth
FROM collars c
LEFT JOIN sample_analyses sa ON c.hole_id = sa.hole_id
LEFT JOIN lithology_logs ll ON c.hole_id = ll.hole_id
GROUP BY c.collar_id;

-- View: Seam Summary by Hole
CREATE VIEW seam_summary_by_hole AS
SELECT 
    sa.hole_id,
    sqc.seam_label as quality_seam_label,
    COUNT(*) as sample_count,
    MIN(sa.depth_from) as seam_top_depth,
    MAX(sa.depth_to) as seam_bottom_depth,
    AVG(sa.ash) as avg_ash,
    AVG(sa.vm) as avg_vm,
    AVG(sa.gross_cv) as avg_gross_cv,
    AVG(sa.net_cv) as avg_net_cv
FROM sample_analyses sa
LEFT JOIN seam_codes_lookup sqc ON sa.seam_quality_id = sqc.seam_id
WHERE sa.seam_quality_id IS NOT NULL
GROUP BY sa.hole_id, sqc.seam_label;

-- =====================================================
-- 5. DOCUMENTATION AND COMMENTS
-- =====================================================

COMMENT ON TABLE collars IS 'Drilling hole collar locations and basic information';
COMMENT ON TABLE lithology_logs IS 'Detailed lithology logs for each depth interval';
COMMENT ON TABLE sample_analyses IS 'Laboratory analysis results for coal samples with foreign key relationships';
COMMENT ON TABLE seam_codes_lookup IS 'Complete lookup table for all seam codes from different classification systems (30, 46, 57, 58, Quality, 73)';
COMMENT ON TABLE rock_codes_lookup IS 'Complete lookup table for all rock/lithology codes';

COMMENT ON COLUMN sample_analyses.seam_quality_id IS 'Foreign key to seam_codes_lookup (Quality classification system)';
COMMENT ON COLUMN sample_analyses.seam_73_id IS 'Foreign key to seam_codes_lookup (73 classification system)';
COMMENT ON COLUMN sample_analyses.depth_from IS 'Starting depth of sample interval (meters)';
COMMENT ON COLUMN sample_analyses.depth_to IS 'Ending depth of sample interval (meters)';
COMMENT ON COLUMN sample_analyses.ash IS 'Ash content percentage (dry basis)';
COMMENT ON COLUMN sample_analyses.vm IS 'Volatile matter percentage (dry ash-free basis)';
COMMENT ON COLUMN sample_analyses.fc IS 'Fixed carbon percentage (dry ash-free basis)';
COMMENT ON COLUMN sample_analyses.gross_cv IS 'Gross Calorific Value in kcal/kg';
COMMENT ON COLUMN sample_analyses.net_cv IS 'Net Calorific Value in kcal/kg';
"""
    
    with open('sql/create_drilling_database_schema.sql', 'w', encoding='utf-8') as f:
        f.write(sql_content)
    
    print("Saved: sql/create_drilling_database_schema.sql")

def create_data_loading_script():
    """Create script to load drilling database data"""
    
    print("\n7. Creating data loading script...")
    
    sql_content = """-- =====================================================
-- Load Coal Drilling Database Data
-- Following Best Practices for Mining Data Management
-- =====================================================

-- Load seam codes lookup
.mode csv
.import data/drilling_database/seam_codes_lookup.csv seam_codes_temp

INSERT INTO seam_codes_lookup (seam_id, system_id, system_name, seam_label, seam_code)
SELECT seam_id, system_id, system_name, seam_label, seam_code
FROM seam_codes_temp
WHERE seam_id != 'seam_id';

DROP TABLE seam_codes_temp;

-- Load rock codes lookup
.import data/drilling_database/rock_codes_lookup.csv rock_codes_temp

INSERT INTO rock_codes_lookup (rock_id, rock_code, lithology, detail)
SELECT rock_id, rock_code, lithology, detail
FROM rock_codes_temp
WHERE rock_id != 'rock_id';

DROP TABLE rock_codes_temp;

-- Load rock types
.import data/drilling_database/rock_types.csv rock_types_temp

INSERT INTO rock_types (rock_type_id, rock_code, rock_type)
SELECT rock_type_id, rock_code, rock_type
FROM rock_types_temp
WHERE rock_type_id != 'rock_type_id';

DROP TABLE rock_types_temp;

-- Load collars
.import data/drilling_database/collars.csv collars_temp

INSERT INTO collars (
    collar_id, hole_id, easting, northing, elevation, 
    azimuth, dip, final_depth, drilling_date, contractor, remarks
)
SELECT 
    collar_id, hole_id, easting, northing, elevation,
    azimuth, dip, final_depth, drilling_date, contractor, remarks
FROM collars_temp
WHERE collar_id != 'collar_id';

DROP TABLE collars_temp;

-- Load lithology logs
.import data/drilling_database/lithology_logs.csv lithology_temp

INSERT INTO lithology_logs (log_id, hole_id, depth_from, depth_to, rock_code, description)
SELECT log_id, hole_id, depth_from, depth_to, rock_code, description
FROM lithology_temp
WHERE log_id != 'log_id';

DROP TABLE lithology_temp;

-- Load sample analyses
.import data/drilling_database/sample_analyses.csv sample_analyses_temp

INSERT INTO sample_analyses (
    sample_id, hole_id, depth_from, depth_to, sample_no,
    im, tm, ash, vm, fc, sulphur, gross_cv, net_cv, sg, rd, hgi,
    seam_code_quality_original, seam_quality_id, seam_73_id,
    analysis_date, lab_name, remarks
)
SELECT 
    sample_id, hole_id, depth_from, depth_to, sample_no,
    im, tm, ash, vm, fc, sulphur, gross_cv, net_cv, sg, rd, hgi,
    seam_code_quality_original,
    CASE WHEN seam_quality_id = '' THEN NULL ELSE seam_quality_id END,
    CASE WHEN seam_73_id = '' THEN NULL ELSE seam_73_id END,
    NULL, NULL, NULL
FROM sample_analyses_temp
WHERE sample_id != 'sample_id';

DROP TABLE sample_analyses_temp;

-- Verify data loading
SELECT 'Seam Codes Count:' as info, COUNT(*) as count FROM seam_codes_lookup
UNION ALL
SELECT 'Rock Codes Count:', COUNT(*) FROM rock_codes_lookup
UNION ALL
SELECT 'Rock Types Count:', COUNT(*) FROM rock_types
UNION ALL
SELECT 'Collars Count:', COUNT(*) FROM collars
UNION ALL
SELECT 'Lithology Logs Count:', COUNT(*) FROM lithology_logs
UNION ALL
SELECT 'Sample Analyses Count:', COUNT(*) FROM sample_analyses;

-- Show sample statistics
SELECT 'Seam Codes by System:' as info;
SELECT system_id, COUNT(*) as count FROM seam_codes_lookup GROUP BY system_id;

SELECT 'Sample Analyses with Seam Codes:' as info;
SELECT COUNT(*) as samples_with_quality FROM sample_analyses WHERE seam_quality_id IS NOT NULL;
SELECT COUNT(*) as samples_with_73 FROM sample_analyses WHERE seam_73_id IS NOT NULL;
"""
    
    with open('sql/load_drilling_database_data.sql', 'w', encoding='utf-8') as f:
        f.write(sql_content)
    
    print("Saved: sql/load_drilling_database_data.sql")

def create_best_practices_documentation():
    """Create documentation for best practices"""
    
    print("\n8. Creating Best Practices documentation...")
    
    doc_content = """# Coal Drilling Database - Best Practices Documentation

## Overview
This database follows best practices for coal drilling data management, ensuring data integrity, normalization, and efficient querying.

## Database Structure

### 1. Lookup Tables
- **seam_codes_lookup**: All seam classification codes from different systems (30, 46, 57, 58, Quality, 73)
- **rock_codes_lookup**: Standard rock/lithology codes
- **rock_types**: Rock type classifications

### 2. Core Tables
- **collars**: Drilling hole collar locations and basic information
- **lithology_logs**: Detailed lithology information for each depth interval
- **sample_analyses_normalized**: Laboratory analysis results with normalized foreign keys

## Key Features

### Normalization
- All seam codes are stored in lookup tables
- Foreign keys ensure referential integrity
- No data duplication
- Easy to maintain and update

### Data Validation
- Depth constraints (depth_to > depth_from)
- Percentage constraints (0-100 for ash, vm, fc, etc.)
- Foreign key constraints ensure data consistency

### Performance Optimization
- Indexes on commonly queried columns:
  - hole_id (primary identifier)
  - depth ranges (for interval queries)
  - seam codes (for filtering)
  - Location coordinates (for spatial queries)

### Views
- `sample_analyses_complete`: Samples with seam information
- `sample_analyses_with_location`: Samples with collar coordinates
- `holes_complete`: Complete hole information with statistics
- `seam_summary_by_hole`: Seam summaries grouped by hole

## Best Practices Applied

### 1. Data Normalization
- ✅ All lookup data in separate tables
- ✅ Foreign keys instead of duplicate data
- ✅ Normalized to 3NF (Third Normal Form)

### 2. Data Integrity
- ✅ Foreign key constraints
- ✅ Check constraints for data validation
- ✅ Unique constraints where appropriate
- ✅ NOT NULL constraints for required fields

### 3. Performance
- ✅ Indexes on foreign keys
- ✅ Indexes on frequently queried columns
- ✅ Composite indexes for common query patterns

### 4. Maintainability
- ✅ Clear table and column names
- ✅ Documentation comments
- ✅ Consistent naming conventions
- ✅ Views for complex queries

### 5. Mining Industry Standards
- ✅ Depth intervals properly tracked
- ✅ Multiple seam classification systems supported
- ✅ Laboratory analysis parameters stored correctly
- ✅ Collar location data included
- ✅ Lithology information linked

## Usage Examples

### Query samples by seam code:
```sql
SELECT * FROM sample_analyses_complete
WHERE quality_seam_label = 'H3c';
```

### Query samples by hole with location:
```sql
SELECT * FROM sample_analyses_with_location
WHERE hole_id = 'BC01C';
```

### Get seam summary:
```sql
SELECT * FROM seam_summary_by_hole
WHERE hole_id = 'BC01C';
```

### Query samples within depth range:
```sql
SELECT * FROM sample_analyses_normalized
WHERE hole_id = 'BC01C'
  AND depth_from >= 4.0
  AND depth_to <= 10.0;
```

## Data Flow

1. **Raw Data**: DH70.xlsx (Excel file)
2. **Lookup Extraction**: Seam codes and rock codes extracted from worksheets
3. **Normalization**: Sample analyses linked to lookup tables via foreign keys
4. **Storage**: Normalized CSV files ready for database import
5. **Database**: SQLite database with proper schema and constraints

## Maintenance

- Update lookup tables when new seam codes or rock codes are added
- Maintain referential integrity when updating sample analyses
- Use transactions for batch updates
- Regular backups of normalized data

## References

- Coal Quality Standards (ASTM, ISO)
- Mining Data Management Best Practices
- Database Normalization Principles
- Spatial Data Management for Mining
"""
    
    os.makedirs('docs', exist_ok=True)
    with open('docs/BEST_PRACTICES.md', 'w', encoding='utf-8') as f:
        f.write(doc_content)
    
    print("Saved: docs/BEST_PRACTICES.md")

def main():
    """Main function"""
    
    # Create drilling database
    sample_analyses_db, seam_codes, rock_codes, collars, lithology_logs, rock_types = create_drilling_database()
    
    # Create SQL schema
    create_drilling_database_schema()
    
    # Create data loading script
    create_data_loading_script()
    
    # Create documentation
    create_best_practices_documentation()
    
    print("\n" + "="*70)
    print("COAL DRILLING DATABASE CREATION COMPLETE!")
    print("="*70)
    print("\nFiles created:")
    print("- data/drilling_database/sample_analyses.csv")
    print("- data/drilling_database/seam_codes_lookup.csv")
    print("- data/drilling_database/rock_codes_lookup.csv")
    print("- data/drilling_database/collars.csv")
    print("- data/drilling_database/lithology_logs.csv")
    print("- data/drilling_database/rock_types.csv")
    print("- sql/create_drilling_database_schema.sql")
    print("- sql/load_drilling_database_data.sql")
    print("- docs/BEST_PRACTICES.md")
    
    return sample_analyses_db, seam_codes, rock_codes

if __name__ == "__main__":
    sample_analyses_db, seam_codes, rock_codes = main()